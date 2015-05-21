#!/usr/bin/env python

import datetime
import MySQLdb as mdb
from excel_db_queries import query_list
from openpyxl import load_workbook
from os.path import dirname, realpath, join
from subprocess import call
from sys import platform
import sys
# import logging

""" This file reads an excel file connects to the mysql database,
    gathers information on hourly impressions delivered, and
    saves it to an excel file (DeliverAllCheck.xlsm).
    
    It is can be run on a server running at least python 2.6, or locally on a mac
    
    The spreadsheet should have an existing tab called 'raw-data'.
    Required files in same directory as this file:
    * my.cnf                        (mysql login credentials)
    * excel_db_queries.py           (queries for each spreadsheet)
    * dropbox_uploader.sh           (for pushing up to Dropbox. Must be configured for dropbox account)
    * DeliverAllCheck-source.xlsm   (This file is not changed)
    
    To Run from command line:
    To run a single command (UpdateTooFast)
    python -c 'import updateDeliverAllCheck; updateDeliverAllCheck.UpdateTooFast()'
    or
    python updateDeliverAllCheck.py allcheck
    python updateDeliverAllCheck.py fast
    python updateDeliverAllCheck.py slow
    
    To run all updates
    python updateDeliverAllCheck.py
"""

# TODO add logger instead of manual logging
# TODO rollover logging
# TODO log to /opt/opinmind/logs/modeling
# TODO add config file for things like wsname?


# get the path of this file (necessary for running from cron job)
parent = dirname(realpath(__file__))
dropboxupdater = join(parent, "./dropbox_uploader.sh")
cnffile = join(parent, "my.cnf")

# =================
# = Set up logger =
# =================
# logging.basicConfig(filename=join(parent, 'update-deliver-log.log'), 
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s')


def IsMac(): return(True if platform=='darwin' else False)

def GetMyCnfGroup():
    """ Return default_group"""
    # change behavior based on platform (mac or server)    
    if platform == 'darwin':
        return "remoteMsql"          # for use with mac
    else:
        return "msql-from-server"    # for use from server (linux)
    
def get_query_results(query):
    """query mysql to get impressions by hour for all campaigns"""
    # con = mdb.connect('localhost', 'testuser', 'test623', 'testdb');
    # http://mysql-python.sourceforge.net/MySQLdb.html
    # db=_mysql.connect(host="outhouse",db="thangs",read_default_file="~/.my.cnf")
    con = mdb.connect(
        read_default_file = cnffile,
        read_default_group = GetMyCnfGroup(),
        db = 'marketplace'
        )
    with con:
        cur = con.cursor()
        cur.execute(query)
        rows = cur.fetchall()
        field_names=[i[0] for i in cur.description]
    return (field_names, rows)

def to_utf(value):
    """ convert to utf-8? """
    if type(value) == str:
        value=unicode(value, "utf-8", errors="ignore")
    return value

def replace_range(query, source_wb_name, target_wb_name, wsname='raw-data'):
    """ update excel workbook with current data, 
        and save to a new file
    """
    # get the workbook
    print log_time(), " loading workbook %s " % source_wb_name
    # logging.info('loading workbook %s: ', source_wb_name)
    wb = load_workbook(filename=source_wb_name, keep_vba=True)
    
    # remove existing 'raw-data' sheet
    ws = wb.get_sheet_by_name(wsname)
    if (ws): wb.remove_sheet(ws)
    
    # print log_time(), " creating new sheet %s " % wsname
    # logging.debug('creating new sheet %s: ', wsname)
    ws = wb.create_sheet(title=wsname)
    
    print log_time(), " querying db"
    # logging.info('querying db')
    # get cursor to query result
    (field_names, rows) = get_query_results(query)
    
    print log_time(), " writing to file"
    # logging.info('writing to file')
    # write query results to 'raw-data'
    ws.append(field_names)      # headers
    for row in rows:            # data
        rowlist = list(row)
        lst = [to_utf(elem) for elem in rowlist]
        ws.append(lst)

    # add timestamp
    ws = wb.get_sheet_by_name('start')
    ws['C4'] = log_time()

    # TODO current version of openpyxl (1.8) does not save to same file name.
    # This bug should be fixed in ver 1.9
    print log_time(), " saving %s" % target_wb_name
    # logging.info('saving %s: ', target_wb_name)
    wb.save(target_wb_name)
    
    print log_time(), "done"
    # logging.info('done')

def log_time():
    """ return current time """
    # TODO Convert to PST
    curtime = datetime.datetime.now()
    if not(IsMac()):
        timetoUTC = datetime.timedelta(hours = 7)
        curtime = curtime - timetoUTC
    # now_pst = now_utc.astimezone('Pacific')
    return str(curtime)

def UpdateOne(query, source, target):
    print log_time(), "begin updating excel", target
    # logging.debug('begin updating excel %s:', target)
    # TODO source != target
    source = join(parent, source)
    target = join(parent, target)
    if source == target:
        print log_time(), "source and target names must be different"
        return
    replace_range(query, source, target)
    
    if not (IsMac()): 
        print "Upload to dropbox"
        # logging.info("Upload to dropbox")
        # print "not mac, not updating right now"
        call([dropboxupdater, "upload", target, "Daily Stats/Regression_Models/."])
    
def UpdateAllCheck():
    """ Update testDeliverAllCheck.xlsm """
    UpdateOne(query_list['deliver_all'], 'DeliverAllCheck-source.xlsm', 'DeliverAllCheck.xlsm')

def UpdateTooFast():
    """ Update testDeliverTooFast.xlsm """
    UpdateOne(query_list['too_fast'], 'DeliverAllCheck-source.xlsm', 'DeliverTooQuickCheck.xlsm')

def UpdateTooSlow():
    """ testDeliverTooSlow.xlsm """
    UpdateOne(query_list['too_slow'], 'DeliverAllCheck-source.xlsm', 'DeliverTooSlowCheck.xlsm')

def main(which='many'):
    """Run all 3 queries, deliver_all, too_fast, too_slow"""
    if   which == 'allcheck':   UpdateAllCheck()
    elif which == 'fast':       UpdateTooFast()
    elif which == 'slow':       UpdateTooSlow()
    else:
        UpdateAllCheck()
        UpdateTooFast()
        UpdateTooSlow()

if __name__ == '__main__':
    # logging.info('BEGIN')
    print('BEGIN')
    which_query = (sys.argv[1] if len(sys.argv) > 1 else 'many')
    main(which_query)
    # logging.info('ALL DONE')
    print('ALL DONE')